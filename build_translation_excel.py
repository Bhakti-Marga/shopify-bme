import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ── colour palette ──────────────────────────────────────────────────────────
CLR_HEADER_BG   = "16254C"   # dark navy  (store brand)
CLR_HEADER_FG   = "FFFFFF"   # white
CLR_SECTION_BG  = "D4AF37"   # gold accent
CLR_SECTION_FG  = "16254C"
CLR_DONE_BG     = "E8F5E9"   # light green  – already Spanish
CLR_ROW_ALT     = "F8F7F4"   # off-white alternate row
CLR_ROW_WHITE   = "FFFFFF"
CLR_BORDER      = "CCCCCC"

thin = Side(style="thin", color=CLR_BORDER)
border = Border(left=thin, right=thin, top=thin, bottom=thin)

def make_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def header_font(bold=True):
    return Font(name="Calibri", bold=bold, color=CLR_HEADER_FG, size=11)

def cell_font(bold=False, color="000000", size=10):
    return Font(name="Calibri", bold=bold, color=color, size=size)

def write_header(ws, titles, row=1):
    for col, title in enumerate(titles, 1):
        c = ws.cell(row=row, column=col, value=title)
        c.font     = Font(name="Calibri", bold=True, color=CLR_HEADER_FG, size=11)
        c.fill     = make_fill(CLR_HEADER_BG)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border   = border

def write_section_row(ws, row_num, label):
    ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=4)
    c = ws.cell(row=row_num, column=1, value=label)
    c.font      = Font(name="Calibri", bold=True, color=CLR_SECTION_FG, size=11)
    c.fill      = make_fill(CLR_SECTION_BG)
    c.alignment = Alignment(horizontal="left", vertical="center")
    c.border    = border

def write_row(ws, row_num, values, alt=False, done=False):
    fill = make_fill(CLR_DONE_BG) if done else make_fill(CLR_ROW_ALT if alt else CLR_ROW_WHITE)
    for col, val in enumerate(values, 1):
        c = ws.cell(row=row_num, column=col, value=val or "")
        c.fill      = fill
        c.font      = cell_font(size=10)
        c.alignment = Alignment(vertical="top", wrap_text=True)
        c.border    = border

# ════════════════════════════════════════════════════════════════════════════
# SHEET 1 — SITEWIDE
# ════════════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Sitewide"
ws1.freeze_panes = "A2"
ws1.column_dimensions["A"].width = 32
ws1.column_dimensions["B"].width = 42
ws1.column_dimensions["C"].width = 42
ws1.column_dimensions["D"].width = 22
ws1.row_dimensions[1].height = 30

COLS = ["Location", "English text", "Spanish translation", "URL slug to change"]
write_header(ws1, COLS)

sitewide_rows = [
    ("Header",              '"Log in"',                              "Iniciar sesión",                          ""),
    ("Header",              '"Start Now" button',                    "Empieza ahora",                           ""),
    ("Footer",              '"All Rights Reserved"',                 "Todos los derechos reservados",           ""),
    ("Cart drawer",         '"Your bag"',                            "Tu bolsa",                                ""),
    ("Cart drawer",         '"Qty:"',                                "Cant.:",                                  ""),
    ("Cart drawer",         '"Order Summary"',                       "Resumen del pedido",                      ""),
    ("Cart drawer",         '"Checkout"',                            "Finalizar compra",                        ""),
    ("Cart drawer",         "empty cart message",                    "Parece que aún no has añadido nada...",   ""),
    ("Search drawer",       '"Search" / "Search..."',                "Buscar",                                  ""),
    ("Mobile nav",          '"Start now"',                           "Empieza ahora",                           ""),
    ("Mobile nav",          '"Log in"',                              "Iniciar sesión",                          ""),
    ("Pagination",          '"Previous" / "Next"',                   "Anterior / Siguiente",                    ""),
    ("Collection filters",  '"Clear Filter" / "Clear All"',          "Limpiar filtro / Limpiar todo",           ""),
    ("Collection filters",  '"From" / "To"',                        "Desde / Hasta",                           ""),
]

for i, row in enumerate(sitewide_rows):
    write_row(ws1, i + 2, row, alt=(i % 2 == 1))

# ════════════════════════════════════════════════════════════════════════════
# SHEET 2 — PAGES
# ════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Pages")
ws2.freeze_panes = "A2"
ws2.column_dimensions["A"].width = 34
ws2.column_dimensions["B"].width = 40
ws2.column_dimensions["C"].width = 40
ws2.column_dimensions["D"].width = 28
ws2.row_dimensions[1].height = 30

write_header(ws2, COLS)

pages_rows = [
    ("/cart",                        '"Your Cart"',                               "Tu carrito",                                ""),
    ("/cart",                        '"remove"',                                  "eliminar",                                  ""),
    ("/cart",                        '"Taxes and shipping calculated at checkout"',"Impuestos y envío calculados al finalizar", ""),
    ("/cart",                        '"continue shopping"',                        "seguir comprando",                          "/collections/all → Shopify system, keep"),
    ("/search",                      '"Products" / "Articles" / "Pages"',          "Productos / Artículos / Páginas",           ""),
    ("/search",                      '"No results found..."',                      "No se encontraron resultados...",           ""),
    ("/search",                      '"EVENT" label',                              "EVENTO",                                    ""),
    ("/collections/[any]",           '"No Products Found"',                        "No se encontraron productos",               ""),
    ("/collections/[any]",           '"Join the Community"',                       "Únete a la comunidad",                      ""),
    ("/collections/[any]",           '"Free Learn to Meditate Course"',            "Curso gratuito Aprende a Meditar",          ""),
    ("/collections/events",          "",                                           "",                                          "events → eventos"),
    ("/products/[donation]",         '"Make this donation every"',                 "Hacer esta donación cada",                  ""),
    ("/products/[donation]",         '"Monthly" / "Quarterly" / "Yearly"',         "Mensual / Trimestral / Anual",              ""),
    ("/products/[donation]",         '"Donate Now"',                               "Donar ahora",                               ""),
    ("/products/make-a-donation",    "",                                           "",                                          "make-a-donation → hacer-una-donacion"),
    ("/products/[event]",            '"Attendees:" / "Qty:"',                      "Asistentes: / Cant.:",                      ""),
    ("/products/[event]",            '"You may be interested in"',                 "También te puede interesar",                ""),
    ("/pages/give",                  '"Give Today"',                               "Dona hoy",                                  "give → dona"),
    ("/pages/all-programs",          '"Discover Events"',                          "Descubre eventos",                          "all-programs → todos-los-programas"),
    ("/pages/all-programs",          '"Soul Awakening Program" / "Bhakti Sundays"',"Programa Despertar del Alma / Domingos Bhakti",""),
    ("/pages/ashram",                '"More info"',                                "Más información",                           "ashram → KEEP (Sanskrit)"),
    ("/pages/ashram",                '"Visit the Temple"',                         "Visita el Templo",                          ""),
    ("/pages/bhakti-sundays",        "",                                           "",                                          "bhakti-sundays → domingos-bhakti"),
    ("/pages/calendar",              "",                                           "",                                          "calendar → calendario"),
    ("/pages/communities",           '"Visit Us"',                                 "Visítanos",                                 "communities → comunidades"),
    ("/pages/communities",           '"Learn More"',                               "Saber más",                                 ""),
    ("/pages/contact",               "",                                           "",                                          "contact → contacto"),
    ("/pages/el-camino",             '"Play Now"',                                 "Reproducir",                                "already Spanish ✓"),
    ("/pages/el-maestro",            '"Play Now"',                                 "Reproducir",                                "already Spanish ✓"),
    ("/pages/mision",                '"Sunday Program" button',                    "Programa del Domingo",                      "already Spanish ✓"),
    ("/pages/om-chanting",           '"Send your name"',                           "Enviar mi nombre",                          "om-chanting → canto-om"),
    ("/pages/project-mantra",        "",                                           "",                                          "project-mantra → proyecto-mantra"),
    ("/pages/request-a-course",      "",                                           "",                                          "request-a-course → solicitar-curso"),
    ("/pages/rituales",              '"Request now" ×3',                           "Solicitar ahora",                           "already Spanish ✓"),
    ("/pages/start-now",             '"Discover Events"',                          "Descubre eventos",                          "start-now → empieza-ahora"),
    ("/pages/start-now",             '"Join the Community" / "Learn to Meditate"', "Únete a la comunidad / Aprende a Meditar",  ""),
    ("/pages/templo",                '"Discover Events"',                          "Descubre eventos",                          "already Spanish ✓"),
    ("/pages/templo",                '"Join Live"',                                "Únete en vivo",                             ""),
    ("/pages/templo",                '"Visit the Temple"',                         "Visita el Templo",                          ""),
    ("/pages/vedic-chanting",        '"Learn more"',                               "Saber más",                                 "vedic-chanting → canto-vedico"),
]

for i, row in enumerate(pages_rows):
    done = "already Spanish" in (row[3] or "")
    write_row(ws2, i + 2, row, alt=(i % 2 == 1), done=done)

# Missing items found after audit
missing_start = len(pages_rows) + 3
write_section_row(ws2, missing_start, "⚠ Found after initial audit — missing translations")
missing_pages = [
    ("/  (homepage hero)",    '"Meet Him" → FIXED ✓',                        "Conócelo",                           "Fixed & pushed ✓"),
    ("event-card.liquid",     'External URL: https://events.bhaktimarga.org/pages/pray-for-the-world — hardcoded English slug on external domain',
                               "Ask bhaktimarga.org team to create /pages/ora-por-el-mundo and update link",      "External domain — manual action needed"),
]
for i, row in enumerate(missing_pages):
    write_row(ws2, missing_start + 1 + i, row, alt=(i % 2 == 1), done=("Fixed" in row[3]))

# ════════════════════════════════════════════════════════════════════════════
# SHEET 3 — BLOG & EVENTS
# ════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Blog & Events")
ws3.freeze_panes = "A2"
ws3.column_dimensions["A"].width = 42
ws3.column_dimensions["B"].width = 40
ws3.column_dimensions["C"].width = 40
ws3.column_dimensions["D"].width = 30
ws3.row_dimensions[1].height = 30

write_header(ws3, COLS)

blog_rows = [
    ("/blogs/events",                                          "",                                      "",                               "events → eventos"),
    ("/blogs/events/[any]",                                    '"Comments"',                             "Comentarios",                    ""),
    ("/blogs/events/darshan",                                  '"Register for Free"',                    "Regístrate gratis",              "darshan → KEEP (Sanskrit)"),
    ("/blogs/events/darshan",                                  '"Buy In-Person Pass"',                   "Compra tu entrada presencial",   ""),
    ("/blogs/events/darshan",                                  '"Contact Us"',                           "Contáctanos",                    ""),
    ("/blogs/events/guruji-birthday",                          '"Get Free General Access Pass"',         "Obtén tu entrada gratuita",      "guruji-birthday → cumpleanos-guruji"),
    ("/blogs/events/guruji-birthday",                          '"Get Premium Access Pass"',              "Obtén tu entrada premium",       ""),
    ("/blogs/events/guruji-birthday",                          '"Contact Us"',                           "Contáctanos",                    ""),
    ("/blogs/events/gurupurnima",                              '"Get Free General Access Pass"',         "Obtén tu entrada gratuita",      "gurupurnima → KEEP (Sanskrit)"),
    ("/blogs/events/gurupurnima",                              '"Get Premium Access Pass"',              "Obtén tu entrada premium",       ""),
    ("/blogs/events/gurupurnima",                              '"Contact Us"',                           "Contáctanos",                    ""),
    ("/blogs/events/krishna-janmashtami",                      '"Get Free General Access Pass"',         "Obtén tu entrada gratuita",      "krishna-janmashtami → KEEP (Sanskrit)"),
    ("/blogs/events/krishna-janmashtami",                      '"Get Premium Access Pass"',              "Obtén tu entrada premium",       ""),
    ("/blogs/events/krishna-janmashtami",                      '"Contact Us"',                           "Contáctanos",                    ""),
    ("/blogs/events/main-event-template",                      '"Buy In-Person Pass"',                   "Compra tu entrada presencial",   "main-event-template → plantilla-evento"),
    ("/blogs/events/main-event-template",                      '"Contact Us"',                           "Contáctanos",                    ""),
    ("/blogs/events/meditation-essentials",                    '"Buy Virtual Pass"',                     "Compra tu entrada virtual",      "meditation-essentials → esenciales-meditacion"),
    ("/blogs/events/meditation-essentials",                    '"Buy In-Person Access Pass"',            "Entrada presencial",             ""),
    ("/blogs/events/meditation-essentials",                    '"Buy In-Person Premium Pass"',           "Entrada presencial premium",     ""),
    ("/blogs/events/meditation-essentials",                    '"Contact Us"',                           "Contáctanos",                    ""),
    ("/blogs/events/narasimha-chaturdashi",                    '"Get Free General Access Pass"',         "Obtén tu entrada gratuita",      "narasimha-chaturdashi → KEEP (Sanskrit)"),
    ("/blogs/events/narasimha-chaturdashi",                    '"Get Premium Access Pass"',              "Obtén tu entrada premium",       ""),
    ("/blogs/events/narasimha-chaturdashi",                    '"Contact Us"',                           "Contáctanos",                    ""),
    ("/blogs/events/navaratri",                                '"Get Free General Access Pass"',         "Obtén tu entrada gratuita",      "navaratri → KEEP (Sanskrit)"),
    ("/blogs/events/navaratri",                                '"Get Premium Access Pass"',              "Obtén tu entrada premium",       ""),
    ("/blogs/events/navaratri",                                '"Contact Us"',                           "Contáctanos",                    ""),
    ("/blogs/events/online-darshan",                           '"Explore" ×3',                           "Explorar",                       "online-darshan → darshan-online"),
    ("/blogs/events/shivaratri",                               '"Get Free General Access Pass"',         "Obtén tu entrada gratuita",      "shivaratri → KEEP (Sanskrit)"),
    ("/blogs/events/shivaratri",                               '"Get Premium Access Pass"',              "Obtén tu entrada premium",       ""),
    ("/blogs/events/shivaratri",                               '"Contact Us"',                           "Contáctanos",                    ""),
    ("/blogs/events/temple-anniversary",                       '"Free General Access Pass"',             "Entrada gratuita",               "temple-anniversary → aniversario-templo"),
    ("/blogs/events/temple-anniversary",                       '"Get Premium Pass"',                     "Entrada premium",                ""),
    ("/blogs/events/temple-anniversary",                       '"Buy In-Person Pass"',                   "Entrada presencial",             ""),
    ("/blogs/events/temple-anniversary",                       '"Contact Us"',                           "Contáctanos",                    ""),
    ("/blogs/events/yogic-philosophy",                         '"Buy In-Person Pass"',                   "Entrada presencial",             "yogic-philosophy → filosofia-yoguica"),
    ("/blogs/events/yogic-philosophy",                         '"Contact Us"',                           "Contáctanos",                    ""),
    ("/blogs/events/4-days-to-soul-awakening-...",             "",                                       "",                               "→ 4-dias-despertar-del-alma-con-paramahamsa-vishwananda"),
]

for i, row in enumerate(blog_rows):
    write_row(ws3, i + 2, row, alt=(i % 2 == 1))

# ════════════════════════════════════════════════════════════════════════════
# SHEET 4 — ACCOUNT
# ════════════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("Account")
ws4.freeze_panes = "A2"
ws4.column_dimensions["A"].width = 28
ws4.column_dimensions["B"].width = 42
ws4.column_dimensions["C"].width = 42
ws4.column_dimensions["D"].width = 30
ws4.row_dimensions[1].height = 30

write_header(ws4, COLS)

account_rows = [
    ("/account/login",    '"Sign In"',                              "Iniciar sesión",              "Shopify system — slug unchangeable"),
    ("/account/login",    '"Create Account"',                       "Crear cuenta",                ""),
    ("/account/login",    '"Reset Password"',                       "Restablecer contraseña",      ""),
    ("/account/login",    '"Email Address"',                        "Correo electrónico",          ""),
    ("/account/login",    '"Password"',                             "Contraseña",                  ""),
    ("/account/login",    '"Forgot Password?"',                     "¿Olvidaste tu contraseña?",   ""),
    ("/account/register", '"Create Account"',                       "Crear cuenta",                ""),
    ("/account/register", '"First Name" / "Last Name"',             "Nombre / Apellido",           ""),
    ("/account/register", "registration description paragraph",     "translate full paragraph",    ""),
    ("/account",          '"My Account"',                           "Mi cuenta",                   ""),
    ("/account",          '"Order History"',                        "Historial de pedidos",        ""),
    ("/account",          '"Manage Addresses"',                     "Gestionar direcciones",       ""),
    ("/account",          '"Sign Out"',                             "Cerrar sesión",               ""),
    ("/account/addresses",'"My Addresses"',                         "Mis direcciones",             ""),
    ("/account/addresses",'"Add New Address"',                      "Añadir nueva dirección",      ""),
    ("/account/addresses","full address form labels",               "translate all fields",        ""),
    ("/account/orders/id",'"Billing Address"',                      "Dirección de facturación",    ""),
    ("/account/orders/id",'"Shipping Address"',                     "Dirección de envío",          ""),
    ("/account/orders/id",'"Subtotal" / "Total" / "Tax" / "Shipping"', "Subtotal / Total / Impuestos / Envío",""),
    ("/account/activate", '"Activate Account"',                     "Activar cuenta",              ""),
    ("/account/activate", '"Decline Invitation"',                   "Rechazar invitación",         ""),
    ("/account/reset",    '"Reset Account Password"',               "Restablecer contraseña",      ""),
]

for i, row in enumerate(account_rows):
    write_row(ws4, i + 2, row, alt=(i % 2 == 1))

# ════════════════════════════════════════════════════════════════════════════
# SHEET 5 — MANUAL CHANGES (Admin URL Handle Checklist)
# ════════════════════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("Manual Changes")
ws5.freeze_panes = "A2"
ws5.column_dimensions["A"].width = 20
ws5.column_dimensions["B"].width = 44
ws5.column_dimensions["C"].width = 44
ws5.column_dimensions["D"].width = 16
ws5.row_dimensions[1].height = 30

CLR_PENDING_BG  = "FFF9E6"   # pale yellow — pending
CLR_DONE2_BG    = "E8F5E9"   # green — done
CLR_STATUS_PEND = "B8860B"   # dark gold text for pending
CLR_STATUS_DONE = "2E7D32"   # green text for done

URL_COLS = ["Section (Admin location)", "Current slug", "New Spanish slug", "Status"]
write_header(ws5, URL_COLS)

# helper: write a URL row
def write_url_row(ws, row_num, values, status="Pending"):
    if status == "Done":
        fill = make_fill(CLR_DONE2_BG)
        status_font = Font(name="Calibri", bold=True, color=CLR_STATUS_DONE, size=10)
    else:
        fill = make_fill(CLR_PENDING_BG)
        status_font = Font(name="Calibri", bold=True, color=CLR_STATUS_PEND, size=10)
    for col, val in enumerate(values, 1):
        c = ws.cell(row=row_num, column=col, value=val or "")
        c.fill      = fill
        c.font      = cell_font(size=10)
        c.alignment = Alignment(vertical="top", wrap_text=True)
        c.border    = border
    # override status cell font
    status_cell = ws.cell(row=row_num, column=4)
    status_cell.font = status_font

# instruction row
ws5.merge_cells(start_row=2, start_column=1, end_row=2, end_column=4)
instr = ws5.cell(row=2, column=1,
    value="How to change a handle: Admin → find item → Search engine listing → Edit → change URL handle → Save → say YES to redirect.")
instr.font      = Font(name="Calibri", italic=True, color="555555", size=10)
instr.fill      = make_fill("F0F0F0")
instr.alignment = Alignment(wrap_text=True, vertical="top")
instr.border    = border
ws5.row_dimensions[2].height = 28

manual_rows = [
    # (section label, current slug, new slug, status)
    # ── PAGES ──
    ("Pages\n(Admin → Online Store → Pages)", "give",          "dona",                    "Pending"),
    ("Pages",                                  "all-programs",  "todos-los-programas",     "Pending"),
    ("Pages",                                  "bhakti-sundays","domingos-bhakti",         "Pending"),
    ("Pages",                                  "calendar",      "calendario",              "Pending"),
    ("Pages",                                  "communities",   "comunidades",             "Pending"),
    ("Pages",                                  "contact",       "contacto",                "Pending"),
    ("Pages",                                  "help-center",   "ayuda",                   "Pending"),
    ("Pages",                                  "om-chanting",   "canto-om",                "Pending"),
    ("Pages",                                  "project-mantra","proyecto-mantra",         "Pending"),
    ("Pages",                                  "request-a-course","solicitar-curso",       "Pending"),
    ("Pages",                                  "start-now",     "empieza-ahora",           "Pending"),
    ("Pages",                                  "vedic-chanting","canto-vedico",            "Pending"),
    # ── PRODUCTS ──
    ("Products\n(Admin → Products)",           "make-a-donation","hacer-una-donacion",     "Pending"),
    # ── BLOG ──
    ("Blog\n(Admin → Blog Posts → Edit Blog)", "events",        "eventos",                 "Pending"),
    # ── ARTICLES ──
    ("Articles\n(Admin → Blog Posts)",         "guruji-birthday",              "cumpleanos-guruji",                                         "Pending"),
    ("Articles",                               "main-event-template",          "plantilla-evento",                                          "Pending"),
    ("Articles",                               "meditation-essentials",        "esenciales-meditacion",                                     "Pending"),
    ("Articles",                               "online-darshan",               "darshan-online",                                            "Pending"),
    ("Articles",                               "temple-anniversary",           "aniversario-templo",                                        "Pending"),
    ("Articles",                               "yogic-philosophy",             "filosofia-yoguica",                                         "Pending"),
    ("Articles",                               "4-days-to-soul-awakening-with-paramahamsa-vishwananda",
                                               "4-dias-despertar-del-alma-con-paramahamsa-vishwananda", "Pending"),
    # ── COLLECTIONS ──
    ("Collections\n(Admin → Products → Collections)", "events", "eventos",                "Pending"),
]

for i, row in enumerate(manual_rows):
    write_url_row(ws5, i + 3, row[:3], status=row[3])

# note row at bottom
last = len(manual_rows) + 4
ws5.merge_cells(start_row=last, start_column=1, end_row=last, end_column=4)
note = ws5.cell(row=last, column=1,
    value="After ALL handle changes are done — push theme: shopify theme push --store d016j0-nz.myshopify.com --password shptka_932058abb3fff3e3c6aed5aaf8f8b69e --theme 183370088792 --allow-live")
note.font      = Font(name="Calibri", italic=True, bold=True, color="16254C", size=10)
note.fill      = make_fill("D4AF37")
note.alignment = Alignment(wrap_text=True, vertical="top")
note.border    = border
ws5.row_dimensions[last].height = 32

# ════════════════════════════════════════════════════════════════════════════
# Save
# ════════════════════════════════════════════════════════════════════════════
output_path = r"C:\Dev\BME\snippets and translation.xlsx"
wb.save(output_path)
print(f"Saved: {output_path}")
